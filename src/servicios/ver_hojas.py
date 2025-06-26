from openpyxl import load_workbook
import os
import sys

# ---- logger --------------------------------------------------------
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.append(ROOT_DIR)
from utils.logger_util import registrar_evento
# -------------------------------------------------------------------

ruta = os.path.join("data", "copias_pc_farmacia", "xFarmaciaJehovaEsMiPastor_20250615-173456.xlsm")
wb = load_workbook(ruta, read_only=True)
print(wb.sheetnames)

registrar_evento(f"Leídas hojas en {os.path.basename(ruta)}")
